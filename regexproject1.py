import re
import openpyxl

teststring = "abcd....????,***"

result = re.findall(r"\,", teststring)
print("this is to test escaping backslash", result)

# testing .,^,$,*,+,?,*?,+?,??,\,[],character-classes,{},|,\A and \Z,\b and \B,\d and \D,\s and \S,\w and \W
# . represents any character except new line # a . represents any character except the new line character i.e. \n , letters upper of lowet, digits , symbos, puncutaliton, commands nad colons and white spaces
# + represents means that the preceding pattern is repeating one or more times
# s matches the space character before the word index

# EXTENSION NOTATION (? extension notation syntax
# non-capturing group ,  adding (?:..) ;  NAMED GROUP  (?P<name> , (result2.groupdict())
# POSITIVE LOOKAHEAD ASSERTIONS (?= ...), NEGATIVE LOOKAHEAD ASSERTION (?!...) # string is applied after
# POSITIVE LOOKBEHIND ASSERTION (?<=...), NEGATIVE LOOKBEHIND ASSERTION (?<!...) # string is applied before

f = open(r"bookshelf.txt")
string = f.read()
print(string)

## match all the authors whose book titles are shorter than 25 characters
# all information  is in same line
# all entries are sepertated by \n
# all the inforation will be seperated by ;

# result1 = re.findall(r".+;(.+\w+);", string , re.M)
result1 = re.findall(r".+?;(.{1,25});", string, re.M)
result2 = re.findall(r".+?;(.{1,25});", string, re.M)

print(result1)
print(result2)

Char_25_books = []
for item in result1:
    if len(item) < 25:
        Char_25_books.append(item)
    else:
        pass
print(Char_25_books)
print(len("Filling the Gap"))

# all the authors who published their books starting with the  year 2000

result3 = re.findall(r"(.+);.+?;20\d{2}", string)
print(result3)

## loading an excel module into python and how to create and handle sheets using openpyxl
# we have two worksheets
# create an object called workbook
workbook = openpyxl.load_workbook(r"Employees.xlsx")  # xlsx is the important expression,
# r is for the raw string  i.e.\n\t will be ignored i.e. ignore special characters
print("""these are the sheets""", workbook.sheetnames)
sheet1 = workbook["EmployeeData"]
dimensions = sheet1.dimensions
print(dimensions)
print(type(workbook.sheetnames))
print(type(sheet1))

for row in sheet1.values:
    print(row)
    # includes the header of the table but all the rows are tuple
    # however we want all the data in this sheet to be represented as a string for re module to work on it

data = []
for row in sheet1.values:  # unpacking each tuple using 7 variables, because we have 7 entries on each row of the table
    a, b, c, d, e, f, g = row  # 7 variables = row which is a tuplel
    data.append("{};{};{};{};{};{};{}".format(a, b, c, d, e, f, g))

# Traceback (most recent call last):
# File "C:/Users/Nitin/PycharmProjects/pythonProject/day-18-start-1/venv/regexproject1.py", line 65, in <module>
#  a,b,c,d,e,f = row # 7 variables = row which is a tuplel
# ValueError: too many values to unpack (expected 6)

print(data)  # this is a list of string where each string contains the data that was previosly part of a tuple
# now we can join the elements of the list using \n as the delimiter
# join function is string.join(iterable)
employees = "\n".join(data)
print(len(employees))
